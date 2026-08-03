"""Extracción de los orígenes Excel: el maestro de lotes vigente y el tareo de personal.

El maestro no es un origen secundario: es la **fuente de identidad de los lotes** (ADR-0003).
Access queda como referencia histórica.
"""

from __future__ import annotations

import csv
import datetime as dt
from pathlib import Path
from typing import Any

from aquanqa_etl.catalogo import (
    MAESTRO_LOTES_COLUMNAS,
    MAESTRO_LOTES_DESTINO,
    MAESTRO_LOTES_FILAS_ESPERADAS,
)
from aquanqa_etl.config import Config
from aquanqa_etl.extract.access import ResultadoExtraccion


def _serializar(valor: Any) -> Any:
    if valor is None:
        return None
    if isinstance(valor, bool):
        return "true" if valor else "false"
    if isinstance(valor, float):
        # openpyxl devuelve todos los números como float; los enteros llegan como 55.0 y
        # escribirlos así obligaría a un cast doble en stg.
        return repr(int(valor)) if valor.is_integer() else repr(valor)
    if isinstance(valor, dt.datetime):
        return valor.isoformat(sep=" ")
    if isinstance(valor, (dt.date, dt.time)):
        return valor.isoformat()
    texto = str(valor).strip()
    return texto if texto else None


def _leer_hoja(ruta: Path, hoja: str | None = None) -> tuple[list[str], list[list[Any]]]:
    from openpyxl import load_workbook

    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        ws = libro[hoja] if hoja else libro[libro.sheetnames[0]]
        filas = ws.iter_rows(values_only=True)
        try:
            cabecera_bruta = next(filas)
        except StopIteration:
            return [], []
        cabecera = [str(c).strip() if c is not None else "" for c in cabecera_bruta]
        datos = [list(f) for f in filas if any(c is not None and str(c).strip() for c in f)]
        return cabecera, datos
    finally:
        libro.close()


def _volcar(
    ruta_csv: Path,
    cols_destino: list[str],
    filas: list[list[Any]],
    indices: list[int],
    con_numero_fila: bool = False,
) -> int:
    ruta_csv.parent.mkdir(parents=True, exist_ok=True)
    with ruta_csv.open("w", encoding="utf-8", newline="") as fh:
        escritor = csv.writer(fh, lineterminator="\n", quoting=csv.QUOTE_NOTNULL)
        escritor.writerow(cols_destino)
        for n, fila in enumerate(filas, start=2):  # 2 = primera fila de datos del Excel
            valores = [_serializar(fila[i]) if i < len(fila) else None for i in indices]
            if con_numero_fila:
                valores.append(str(n))
            escritor.writerow(valores)
    return len(filas)


def extraer_maestro_lotes(config: Config, registrar=print) -> ResultadoExtraccion:
    ruta = config.maestro_lotes
    if not ruta.exists():
        raise FileNotFoundError(
            f"No encuentro el maestro de lotes en {ruta}.\n"
            "  Es la fuente de identidad de los lotes: sin él no se puede resolver ningún "
            "hecho (ADR-0003). Revisa MAESTRO_LOTES_PATH en .env."
        )

    inicio = dt.datetime.now(dt.UTC)
    cabecera, filas = _leer_hoja(ruta)

    faltan = [o for o, _ in MAESTRO_LOTES_COLUMNAS if o not in cabecera]
    if faltan:
        raise ValueError(
            f"El maestro {ruta.name} no tiene las columnas {faltan}.\n"
            f"  Encontradas: {cabecera}\n"
            "  Si el maestro cambió de estructura, hay que actualizar MAESTRO_LOTES_COLUMNAS "
            "en catalogo.py antes de cargar: cargar a ciegas rompería la identidad de lote."
        )

    indices = [cabecera.index(o) for o, _ in MAESTRO_LOTES_COLUMNAS]
    destino = config.csv_de(MAESTRO_LOTES_DESTINO)
    n = _volcar(destino, [d for _, d in MAESTRO_LOTES_COLUMNAS], filas, indices)

    resultado = ResultadoExtraccion(
        tabla=ruta.name,
        destino=MAESTRO_LOTES_DESTINO,
        filas=n,
        ruta_csv=destino,
        esperadas=MAESTRO_LOTES_FILAS_ESPERADAS,
        extraido_en=inicio,
    )
    marca = "ok" if resultado.ok else f"DESVIACIÓN {resultado.desviacion:+d}"
    registrar(f"  {ruta.name:<24} {n:>8,} filas  {marca}")
    return resultado


def extraer_tareo(config: Config, registrar=print) -> ResultadoExtraccion | None:
    """Extrae el tareo si está disponible. Devuelve None si no lo está.

    Es opcional a propósito: sin el tareo todo migra salvo el dominio de personal, y bloquear
    la migración completa por un archivo que vive en una biblioteca SharePoint ajena sería
    desproporcionado.
    """
    ruta = config.tareo
    if not ruta.exists():
        registrar(
            f"  tareo · no disponible en {ruta}\n"
            "      El informe SEGUIMIENTO DE PERSONAL depende de este archivo (B-1/B-5).\n"
            "      Sin él se migra todo excepto el dominio de personal."
        )
        return None

    inicio = dt.datetime.now(dt.UTC)
    cabecera, filas = _leer_hoja(ruta)
    registrar(f"  tareo · columnas encontradas: {cabecera}")

    # El mapeo se resuelve por coincidencia laxa porque la estructura del archivo aún no se ha
    # podido inspeccionar; las columnas que no se reconozcan se reportan en lugar de perderse.
    alias = {
        "documento": ("documento", "dni", "nro documento", "n documento"),
        "nombre": ("nombre", "nombres", "trabajador", "apellidos y nombres"),
        "fecha": ("fecha", "dia", "día"),
        "horas": ("horas", "hrs", "hh", "horas trabajadas"),
        "labor": ("labor", "actividad", "tarea"),
        "fundo": ("fundo", "empresa"),
        "modulo": ("modulo", "módulo"),
        "lote": ("lote",),
    }
    normalizada = [c.strip().lower() for c in cabecera]
    indices: list[int] = []
    destino_cols: list[str] = []
    no_encontradas: list[str] = []
    for destino_col, candidatos in alias.items():
        idx = next((normalizada.index(c) for c in candidatos if c in normalizada), None)
        if idx is None:
            no_encontradas.append(destino_col)
            continue
        indices.append(idx)
        destino_cols.append(destino_col)

    if no_encontradas:
        registrar(f"      columnas no reconocidas y por tanto vacías: {no_encontradas}")
    if "documento" not in destino_cols:
        raise ValueError(
            f"El tareo {ruta.name} no tiene una columna de documento reconocible "
            f"(buscadas: {alias['documento']}).\n"
            "  Sin DNI no hay forma de enlazarlo con las evaluaciones (H-09)."
        )

    destino_csv = config.csv_de("tareo")
    n = _volcar(destino_csv, [*destino_cols, "origen_fila"], filas, indices, con_numero_fila=True)
    registrar(f"  {ruta.name:<24} {n:>8,} filas  ok")
    return ResultadoExtraccion(
        tabla=ruta.name,
        destino="tareo",
        filas=n,
        ruta_csv=destino_csv,
        esperadas=None,
        extraido_en=inicio,
    )
